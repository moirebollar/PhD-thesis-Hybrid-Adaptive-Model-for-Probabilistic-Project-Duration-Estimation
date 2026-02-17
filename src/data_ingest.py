"""
Data Ingestion & Normalization Module.

Reads Gantt/Activity list Excel files from historical projects,
normalizes activities to standardized categories, applies scaling
factors based on project capacity, and outputs a canonical DataFrame.
"""

import json
import warnings
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd


# Standardized activity categories for normalization
STANDARD_CATEGORIES = [
    "design",
    "permitting",
    "site_preparation",
    "civil_works",
    "electromechanical",
    "piping",
    "instrumentation",
    "electrical",
    "testing",
    "commissioning",
    "startup",
]

# Keywords to auto-classify activities (Spanish & English)
CATEGORY_KEYWORDS = {
    "design": [
        "diseño", "design", "ingeniería", "engineering", "plano", "drawing",
        "cálculo", "calculation", "especificación", "specification",
    ],
    "permitting": [
        "permiso", "permit", "licencia", "license", "trámite", "regulat",
        "aprobación", "approval", "autorización", "authorization",
    ],
    "site_preparation": [
        "preparación", "preparation", "terreno", "site", "excavación",
        "excavation", "cimentación", "foundation", "demolición", "demolition",
    ],
    "civil_works": [
        "civil", "concreto", "concrete", "estructura", "structure",
        "albañilería", "masonry", "obra civil", "construcción", "construction",
    ],
    "electromechanical": [
        "electromecánic", "electromechanical", "equipo", "equipment",
        "montaje", "assembly", "instalación mecánica", "mechanical",
        "bomba", "pump", "motor", "reactor", "tanque", "tank",
    ],
    "piping": [
        "tubería", "piping", "pipe", "válvula", "valve", "conexión",
        "connection", "manifold", "header",
    ],
    "instrumentation": [
        "instrumentación", "instrumentation", "sensor", "control",
        "PLC", "SCADA", "medición", "measurement", "calibración",
    ],
    "electrical": [
        "eléctric", "electric", "cableado", "wiring", "tablero", "panel",
        "transformador", "transformer", "acometida", "power supply",
    ],
    "testing": [
        "prueba", "test", "verificación", "verification", "inspección",
        "inspection", "protocolo", "protocol", "FAT", "SAT",
    ],
    "commissioning": [
        "puesta en marcha", "commissioning", "arranque", "startup",
        "operación", "operation", "ajuste", "tuning", "optimización",
    ],
    "startup": [
        "entrega", "handover", "capacitación", "training", "garantía",
        "warranty", "cierre", "closeout",
    ],
}


def load_config(config_path: Optional[str] = None) -> dict:
    """Load configuration from config.json."""
    if config_path is None:
        config_path = Path(__file__).parent.parent / "config.json"
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def classify_activity(activity_name: str) -> str:
    """Classify an activity into a standard category based on keywords."""
    name_lower = activity_name.lower().strip()
    best_match = "electromechanical"  # default category
    best_score = 0

    for category, keywords in CATEGORY_KEYWORDS.items():
        score = sum(1 for kw in keywords if kw.lower() in name_lower)
        if score > best_score:
            best_score = score
            best_match = category

    return best_match


def read_activity_excel(
    filepath: str,
    sheet_name: Optional[str] = None,
    column_mapping: Optional[dict] = None,
) -> pd.DataFrame:
    """
    Read an activity list Excel file and return a standardized DataFrame.

    Parameters
    ----------
    filepath : str
        Path to the Excel file.
    sheet_name : str, optional
        Sheet name to read. If None, reads the first sheet.
    column_mapping : dict, optional
        Mapping from actual column names to standard names.
        Standard names: 'id', 'name', 'wbs_code', 'predecessors',
        'duration', 'optimistic', 'most_likely', 'pessimistic',
        'start_date', 'end_date', 'category'.

    Returns
    -------
    pd.DataFrame
        Standardized activity DataFrame.
    """
    df = pd.read_excel(filepath, sheet_name=sheet_name or 0, engine="openpyxl")

    # Apply column mapping if provided
    if column_mapping:
        df = df.rename(columns=column_mapping)

    # Auto-detect column names (Spanish/English)
    auto_map = _auto_detect_columns(df.columns.tolist())
    df = df.rename(columns=auto_map)

    # Ensure required columns exist
    if "name" not in df.columns:
        raise ValueError(
            f"Could not identify activity name column in {filepath}. "
            f"Columns found: {df.columns.tolist()}"
        )

    # Generate ID if missing
    if "id" not in df.columns:
        df["id"] = range(1, len(df) + 1)

    # Auto-classify activities if category not provided
    if "category" not in df.columns:
        df["category"] = df["name"].apply(classify_activity)

    # Calculate duration from dates if not directly provided
    if "duration" not in df.columns and "start_date" in df.columns and "end_date" in df.columns:
        df["start_date"] = pd.to_datetime(df["start_date"])
        df["end_date"] = pd.to_datetime(df["end_date"])
        df["duration"] = (df["end_date"] - df["start_date"]).dt.days

    # Parse predecessors (handle various formats: "1,2,3" or "1;2;3" or "1FS,2FS")
    if "predecessors" in df.columns:
        df["predecessors"] = df["predecessors"].apply(_parse_predecessors)
    else:
        df["predecessors"] = [[] for _ in range(len(df))]

    # Set distribution type default
    if "dist_type" not in df.columns:
        df["dist_type"] = "betapert"

    # Generate three-point estimates from duration if not provided
    if "most_likely" not in df.columns and "duration" in df.columns:
        df["most_likely"] = df["duration"].astype(float)
        df["optimistic"] = (df["most_likely"] * 0.75).round(1)
        df["pessimistic"] = (df["most_likely"] * 1.40).round(1)
    elif "most_likely" in df.columns:
        df["optimistic"] = df.get("optimistic", df["most_likely"] * 0.75)
        df["pessimistic"] = df.get("pessimistic", df["most_likely"] * 1.40)

    # Rename three-point estimate columns to standard (a, m, b)
    rename_amb = {}
    if "optimistic" in df.columns:
        rename_amb["optimistic"] = "a"
    if "most_likely" in df.columns:
        rename_amb["most_likely"] = "m"
    if "pessimistic" in df.columns:
        rename_amb["pessimistic"] = "b"
    df = df.rename(columns=rename_amb)

    # Ensure numeric types
    for col in ["a", "m", "b"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Validate a <= m <= b
    if all(c in df.columns for c in ["a", "m", "b"]):
        invalid = df[(df["a"] > df["m"]) | (df["m"] > df["b"])]
        if len(invalid) > 0:
            warnings.warn(
                f"Found {len(invalid)} activities where a > m or m > b. "
                "Auto-correcting by sorting values."
            )
            for idx in invalid.index:
                vals = sorted([df.at[idx, "a"], df.at[idx, "m"], df.at[idx, "b"]])
                df.at[idx, "a"] = vals[0]
                df.at[idx, "m"] = vals[1]
                df.at[idx, "b"] = vals[2]

    return df


def _auto_detect_columns(columns: list) -> dict:
    """Auto-detect and map column names to standard names."""
    mapping = {}
    col_lower = {c: c.lower().strip() for c in columns}

    patterns = {
        "id": ["id", "num", "número", "numero", "no.", "#", "act_id"],
        "name": ["nombre", "name", "actividad", "activity", "descripción",
                 "description", "tarea", "task"],
        "wbs_code": ["wbs", "edt", "código", "code"],
        "predecessors": ["predecesoras", "predecessors", "dependencias",
                        "dependencies", "pred"],
        "duration": ["duración", "duration", "días", "days", "tiempo", "time"],
        "optimistic": ["optimista", "optimistic", "min", "a", "t_opt"],
        "most_likely": ["más probable", "most likely", "modal", "m", "t_ml",
                       "más_probable"],
        "pessimistic": ["pesimista", "pessimistic", "max", "b", "t_pes"],
        "start_date": ["inicio", "start", "fecha_inicio", "start_date"],
        "end_date": ["fin", "end", "finish", "fecha_fin", "end_date"],
        "category": ["categoría", "category", "tipo", "type", "fase", "phase"],
    }

    for std_name, keywords in patterns.items():
        for orig_col, lower_col in col_lower.items():
            if any(kw in lower_col for kw in keywords):
                if std_name not in mapping.values():
                    mapping[orig_col] = std_name
                    break

    return mapping


def _parse_predecessors(value) -> list:
    """Parse predecessor field into a list of activity IDs."""
    if pd.isna(value) or value == "" or value is None:
        return []

    value_str = str(value).strip()
    if not value_str:
        return []

    # Remove dependency type suffixes (FS, SS, FF, SF)
    import re
    parts = re.split(r"[,;\s]+", value_str)
    result = []
    for part in parts:
        # Extract numeric ID, ignoring FS/SS/FF/SF suffixes
        cleaned = re.sub(r"[A-Za-z+\-]+.*$", "", part.strip())
        if cleaned:
            try:
                result.append(int(float(cleaned)))
            except ValueError:
                continue
    return result


def normalize_projects(
    project_dfs: list[tuple[str, pd.DataFrame, float]],
    reference_capacity: Optional[float] = None,
) -> pd.DataFrame:
    """
    Normalize multiple projects to a common scale.

    Parameters
    ----------
    project_dfs : list of (project_id, DataFrame, capacity)
        Each tuple contains the project identifier, its activity DataFrame,
        and the project capacity (e.g., m³/day of water treatment).
    reference_capacity : float, optional
        Reference capacity for scaling. If None, uses the median capacity.

    Returns
    -------
    pd.DataFrame
        Combined DataFrame with all projects normalized, including
        'project_id', 'scale_factor', and 'normalized_duration' columns.
    """
    capacities = [cap for _, _, cap in project_dfs]
    if reference_capacity is None:
        reference_capacity = float(np.median(capacities))

    all_dfs = []
    for project_id, df, capacity in project_dfs:
        df = df.copy()
        df["project_id"] = project_id

        # Scaling factor: power-law relationship (economies of scale)
        # scale_factor = (reference_capacity / actual_capacity) ^ exponent
        # Exponent ~0.6 is typical for chemical/industrial plants
        scale_exponent = 0.6
        scale_factor = (reference_capacity / capacity) ** scale_exponent
        df["scale_factor"] = scale_factor

        # Apply scaling to duration estimates
        for col in ["a", "m", "b"]:
            if col in df.columns:
                df[f"{col}_original"] = df[col]
                df[col] = (df[col] * scale_factor).round(1)

        all_dfs.append(df)

    combined = pd.concat(all_dfs, ignore_index=True)
    return combined


def convert_time_units(
    df: pd.DataFrame,
    from_unit: str = "calendar_days",
    to_unit: str = "calendar_days",
    working_days_per_week: int = 5,
) -> pd.DataFrame:
    """
    Convert time units between calendar days and working days.

    Parameters
    ----------
    df : pd.DataFrame
        Activity DataFrame with columns 'a', 'm', 'b'.
    from_unit : str
        Current unit: 'calendar_days' or 'working_days'.
    to_unit : str
        Target unit: 'calendar_days' or 'working_days'.
    working_days_per_week : int
        Number of working days per week (default: 5).

    Returns
    -------
    pd.DataFrame
        DataFrame with converted duration columns.
    """
    if from_unit == to_unit:
        return df

    df = df.copy()
    factor = 7.0 / working_days_per_week  # calendar / working ratio

    for col in ["a", "m", "b"]:
        if col not in df.columns:
            continue
        if from_unit == "working_days" and to_unit == "calendar_days":
            df[col] = (df[col] * factor).round(1)
        elif from_unit == "calendar_days" and to_unit == "working_days":
            df[col] = (df[col] / factor).round(1)

    return df


def interpolate_missing(df: pd.DataFrame) -> pd.DataFrame:
    """
    Handle missing data in activity estimates by interpolation.

    Strategy:
    - If 'm' is missing but 'a' and 'b' exist: m = (a + b) / 2
    - If 'a' is missing: a = m * 0.75
    - If 'b' is missing: b = m * 1.40
    - If all three are missing and 'duration' exists: derive from duration
    """
    df = df.copy()

    for idx in df.index:
        a = df.at[idx, "a"] if "a" in df.columns and pd.notna(df.at[idx, "a"]) else None
        m = df.at[idx, "m"] if "m" in df.columns and pd.notna(df.at[idx, "m"]) else None
        b = df.at[idx, "b"] if "b" in df.columns and pd.notna(df.at[idx, "b"]) else None

        if m is None and a is not None and b is not None:
            df.at[idx, "m"] = round((a + b) / 2, 1)
        elif m is None and "duration" in df.columns and pd.notna(df.at[idx, "duration"]):
            df.at[idx, "m"] = float(df.at[idx, "duration"])
            m = df.at[idx, "m"]

        m = df.at[idx, "m"] if "m" in df.columns and pd.notna(df.at[idx, "m"]) else None

        if a is None and m is not None:
            df.at[idx, "a"] = round(m * 0.75, 1)
        if b is None and m is not None:
            df.at[idx, "b"] = round(m * 1.40, 1)

    return df


def build_canonical_dataframe(
    filepath_or_df,
    project_id: str = "P1",
    capacity: float = 1.0,
    column_mapping: Optional[dict] = None,
    time_unit: str = "calendar_days",
) -> pd.DataFrame:
    """
    One-stop function: load, clean, classify, interpolate, and return
    a canonical activity DataFrame ready for simulation.

    Parameters
    ----------
    filepath_or_df : str or pd.DataFrame
        Path to Excel file or pre-loaded DataFrame.
    project_id : str
        Project identifier.
    capacity : float
        Project capacity for scaling.
    column_mapping : dict, optional
        Column name mapping.
    time_unit : str
        Time unit of the input data.

    Returns
    -------
    pd.DataFrame
        Canonical DataFrame with columns:
        id, name, wbs_code, predecessors, dist_type, a, m, b, category, project_id
    """
    if isinstance(filepath_or_df, (str, Path)):
        df = read_activity_excel(str(filepath_or_df), column_mapping=column_mapping)
    else:
        df = filepath_or_df.copy()

    df = interpolate_missing(df)

    config = load_config()
    target_unit = config.get("time_units", {}).get("default", "calendar_days")
    working_days = config.get("time_units", {}).get("working_days_per_week", 5)
    df = convert_time_units(df, from_unit=time_unit, to_unit=target_unit,
                            working_days_per_week=working_days)

    df["project_id"] = project_id

    # Ensure all required columns exist
    required = ["id", "name", "predecessors", "dist_type", "a", "m", "b", "category"]
    for col in required:
        if col not in df.columns:
            if col == "wbs_code":
                df[col] = ""
            elif col == "dist_type":
                df[col] = "betapert"
            elif col == "predecessors":
                df[col] = [[] for _ in range(len(df))]

    return df


def create_activity_template(output_path: str) -> None:
    """Create an Excel template for activity data entry."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Activities"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        ("ID", 8),
        ("WBS Code", 12),
        ("Activity Name", 40),
        ("Predecessors\n(comma-separated IDs)", 20),
        ("Distribution Type\n(triangular/betapert)", 18),
        ("Optimistic\n(days)", 14),
        ("Most Likely\n(days)", 14),
        ("Pessimistic\n(days)", 14),
        ("Category", 18),
        ("Notes", 30),
    ]

    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = width

    # Example rows
    examples = [
        (1, "1.1", "Ingeniería básica y de detalle", "", "betapert", 20, 30, 45, "design", ""),
        (2, "1.2", "Gestión de permisos", "1", "betapert", 15, 25, 50, "permitting", ""),
        (3, "2.1", "Preparación del terreno", "1", "betapert", 5, 10, 18, "site_preparation", ""),
        (4, "2.2", "Obra civil y cimentaciones", "3", "betapert", 20, 30, 50, "civil_works", ""),
        (5, "3.1", "Montaje electromecánico", "2,4", "betapert", 25, 40, 65, "electromechanical", ""),
        (6, "3.2", "Instalación de tuberías", "4", "betapert", 15, 22, 35, "piping", ""),
        (7, "3.3", "Instrumentación y control", "5,6", "betapert", 10, 15, 25, "instrumentation", ""),
        (8, "4.1", "Pruebas y verificación", "7", "betapert", 8, 12, 22, "testing", ""),
        (9, "4.2", "Puesta en marcha", "8", "betapert", 10, 18, 35, "commissioning", ""),
        (10, "4.3", "Entrega y cierre", "9", "betapert", 5, 7, 14, "startup", ""),
    ]

    light_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = light_fill

    # Instructions sheet
    ws_inst = wb.create_sheet("Instructions")
    instructions = [
        "ACTIVITY TEMPLATE - INSTRUCTIONS",
        "",
        "This template is for entering project activity data for Monte Carlo simulation.",
        "",
        "COLUMN DESCRIPTIONS:",
        "- ID: Unique numeric identifier for each activity",
        "- WBS Code: Work Breakdown Structure code (optional)",
        "- Activity Name: Descriptive name of the activity",
        "- Predecessors: Comma-separated IDs of activities that must finish before this one starts",
        "- Distribution Type: 'triangular' or 'betapert' (recommended: betapert)",
        "- Optimistic (days): Best-case duration estimate (a)",
        "- Most Likely (days): Most probable duration estimate (m)",
        "- Pessimistic (days): Worst-case duration estimate (b)",
        "- Category: Activity category for classification and analysis",
        "",
        "VALID CATEGORIES:",
        "  design, permitting, site_preparation, civil_works, electromechanical,",
        "  piping, instrumentation, electrical, testing, commissioning, startup",
        "",
        "NOTES:",
        "- Optimistic <= Most Likely <= Pessimistic must always hold",
        "- BetaPERT distribution uses lambda=4 weighting on the most likely value",
        "- Predecessors define Finish-to-Start relationships",
        "- All durations should be in the same time unit (calendar days or working days)",
    ]
    for row_idx, line in enumerate(instructions, 1):
        cell = ws_inst.cell(row=row_idx, column=1, value=line)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif line.startswith("- ") or line.startswith("  "):
            cell.font = Font(size=10)
    ws_inst.column_dimensions["A"].width = 80

    wb.save(output_path)


def create_expert_questionnaire(output_path: str) -> None:
    """Create an Excel questionnaire for Delphi expert data collection."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()

    # --- Sheet 1: Expert Info ---
    ws1 = wb.active
    ws1.title = "Expert Information"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    info_fields = [
        "Expert ID (assigned)",
        "Name (optional)",
        "Years of experience in industrial water treatment",
        "Number of similar projects completed",
        "Current role / specialization",
        "Date of questionnaire completion",
    ]
    for row_idx, field in enumerate(info_fields, 1):
        cell_label = ws1.cell(row=row_idx, column=1, value=field)
        cell_label.font = Font(bold=True)
        ws1.cell(row=row_idx, column=2, value="")
        ws1.column_dimensions["A"].width = 50
        ws1.column_dimensions["B"].width = 30

    # --- Sheet 2: Duration Estimates ---
    ws2 = wb.create_sheet("Duration Estimates")
    duration_headers = [
        "Activity ID", "Activity Name", "Optimistic (days)",
        "Most Likely (days)", "Pessimistic (days)", "Confidence (1-5)",
        "Comments",
    ]
    for col_idx, header in enumerate(duration_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
    ws2.column_dimensions["B"].width = 35
    ws2.column_dimensions["G"].width = 30

    # Confidence validation (1-5)
    dv_conf = DataValidation(type="whole", operator="between", formula1=1, formula2=5)
    dv_conf.error = "Confidence must be between 1 and 5"
    dv_conf.errorTitle = "Invalid confidence"
    ws2.add_data_validation(dv_conf)
    dv_conf.add("F2:F100")

    # --- Sheet 3: Uncertainty Sources ---
    ws3 = wb.create_sheet("Uncertainty Sources")
    unc_headers = [
        "Variable ID", "Uncertainty Variable",
        "Impact Rating (1-5)", "Likelihood Rating (1-5)",
        "Affected Activities (comma-separated IDs)",
        "Description / Justification",
    ]
    for col_idx, header in enumerate(unc_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Pre-fill uncertainty variables from thesis
    uncertainty_vars = [
        ("U1", "Atypical contaminant composition"),
        ("U2", "Regulatory changes during project"),
        ("U3", "Equipment vendor delays"),
        ("U4", "Weather / environmental conditions"),
        ("U5", "Technology performance uncertainty"),
        ("U6", "Skilled labor availability"),
        ("U7", "Design changes / scope creep"),
        ("U8", "Permitting delays"),
        ("U9", "Site access restrictions"),
        ("U10", "Supply chain disruptions"),
    ]
    for row_idx, (var_id, var_name) in enumerate(uncertainty_vars, 2):
        ws3.cell(row=row_idx, column=1, value=var_id)
        ws3.cell(row=row_idx, column=2, value=var_name)
    ws3.column_dimensions["B"].width = 35
    ws3.column_dimensions["F"].width = 40

    # --- Sheet 4: Risk Identification ---
    ws4 = wb.create_sheet("Risk Identification")
    risk_headers = [
        "Risk ID", "Risk Description",
        "Probability (%)", "Impact - Min (days)",
        "Impact - Most Likely (days)", "Impact - Max (days)",
        "Affected Activities", "Mitigation Strategy",
    ]
    for col_idx, header in enumerate(risk_headers, 1):
        cell = ws4.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
    ws4.column_dimensions["B"].width = 40
    ws4.column_dimensions["H"].width = 40

    # --- Sheet 5: Instructions ---
    ws5 = wb.create_sheet("Instructions")
    instructions = [
        "EXPERT QUESTIONNAIRE - DELPHI METHOD",
        "",
        "This questionnaire is part of a doctoral research on probabilistic estimation",
        "of project durations for water recovery plant construction.",
        "",
        "INSTRUCTIONS:",
        "1. Fill in your information on the 'Expert Information' sheet",
        "2. Provide three-point duration estimates for each activity on 'Duration Estimates'",
        "   - Optimistic: minimum time under ideal conditions",
        "   - Most Likely: most probable duration based on your experience",
        "   - Pessimistic: maximum time under adverse conditions",
        "   - Confidence (1-5): how confident you are in your estimate",
        "3. Rate uncertainty sources on 'Uncertainty Sources'",
        "   - Impact (1-5): how much this variable affects project duration",
        "   - Likelihood (1-5): how likely this variable is to occur",
        "4. Identify additional risks on 'Risk Identification'",
        "",
        "All information will be treated confidentially.",
        "Thank you for your participation.",
    ]
    for row_idx, line in enumerate(instructions, 1):
        cell = ws5.cell(row=row_idx, column=1, value=line)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
    ws5.column_dimensions["A"].width = 80

    wb.save(output_path)


def create_risk_registry_template(output_path: str) -> None:
    """Create an Excel template for risk registry data entry."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Risk Registry"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        ("Risk ID", 10),
        ("Risk Name", 35),
        ("Description", 45),
        ("Probability (0-1)", 16),
        ("Impact Distribution\n(triangular/betapert/uniform)", 20),
        ("Impact Min (days)", 16),
        ("Impact Most Likely (days)", 20),
        ("Impact Max (days)", 16),
        ("Affected Activities\n(comma-separated IDs or 'all')", 25),
        ("Risk Category", 18),
        ("Mitigation Available", 18),
        ("Status (active/mitigated/closed)", 16),
    ]

    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        col_letter = chr(64 + col_idx) if col_idx <= 26 else f"A{chr(64 + col_idx - 26)}"
        ws.column_dimensions[col_letter].width = width

    # Example risks from thesis
    examples = [
        ("R1", "Unexpected contaminant", "Atypical contaminant discovered requiring process redesign",
         0.25, "triangular", 8, 15, 28, "1,2,5", "technical", "No", "active"),
        ("R2", "Extreme weather event", "Severe weather causing construction delays",
         0.20, "triangular", 5, 10, 20, "3,4,5,6", "environmental", "No", "active"),
        ("R3", "Equipment delivery delay", "Late delivery of critical equipment from vendor",
         0.30, "betapert", 10, 20, 45, "5,6", "supply_chain", "Yes", "active"),
        ("R4", "Regulatory change", "New regulatory requirements during construction",
         0.15, "triangular", 15, 30, 60, "1,2", "regulatory", "No", "active"),
        ("R5", "Unidentified risk reserve", "Unknown-unknown risk events",
         0.07, "uniform", 15, 45, 90, "all", "residual", "No", "active"),
    ]

    light_fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = light_fill

    wb.save(output_path)


def generate_all_templates(output_dir: str) -> None:
    """Generate all Excel templates in the specified directory."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    create_activity_template(str(output_dir / "activity_template.xlsx"))
    create_expert_questionnaire(str(output_dir / "expert_questionnaire.xlsx"))
    create_risk_registry_template(str(output_dir / "risk_registry.xlsx"))
    print(f"All templates generated in: {output_dir}")
